"""
Profile all 16 historical MA files against Feb 2026 canonical reference.
Outputs:
  - profile_report.md  (human-readable diff report)
  - profile_detail.csv (per-file per-sheet details)
  - canonical_schema.json (authoritative Feb 2026 structure)
"""
import os
import json
import csv
import openpyxl
from datetime import datetime, date
from collections import OrderedDict

UPLOADS_DIR = "/sessions/dazzling-upbeat-hopper/mnt/uploads"
OUT_DIR = "/sessions/dazzling-upbeat-hopper"

# Chronological order (oldest -> newest). Canonical = last entry.
FILES = [
    ("2024-11", "FY25 Management Accounts - November 24.xlsx"),
    ("2024-12", "FY25 Management Accounts - December 24.xlsx"),
    ("2025-01", "FY25 Management Accounts - January 25.xlsx"),
    ("2025-02", "FY25 Management Accounts - February 25.xlsx"),
    ("2025-03", "FY25 Management Accounts - March 25 (1).xlsx"),
    ("2025-04", "FY25 Management Accounts - April 25.xlsx"),
    ("2025-05", "FY25 Management Accounts - May 25.xlsx"),
    ("2025-06", "FY25 Management Accounts - June 25.xlsx"),
    ("2025-07", "FY25 Management Accounts - July 25.xlsx"),
    ("2025-08", "FY25 Management Accounts - August 25.xlsx"),
    ("2025-09", "FY25 Management Accounts - September 25.xlsx"),
    ("2025-10", "FY25 Management Accounts - October 25.xlsx"),
    ("2025-11", "FY26 Management Accounts - November 25.xlsx"),
    ("2025-12", "2. FY26 Management Accounts - December 25.xlsx"),
    ("2026-01", "3. FY26 Management Accounts - January 26.xlsx"),
    ("2026-02", "4. FY26 Management Accounts - February 26.xlsx"),  # <-- CANONICAL
]

CANONICAL_KEY = "2026-02"

# Sheets we care about for the pipeline
CORE_SHEETS = [
    "P&L Detail",
    "P&L Summary",
    "Financial KPIs",
    "Balance Sheet",
    "Headcount",
    "Revenue Waterfall",
]

def safe_str(v):
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()

def sheet_lookup(wb, target):
    """Return actual sheet name (handling trailing spaces / case drift)."""
    target_n = target.strip().lower()
    for s in wb.sheetnames:
        if s.strip().lower() == target_n:
            return s
    return None

def profile_sheet(ws, max_rows=60, max_cols=8):
    """Capture a skeleton of a sheet: labels in col A/B and first-row values."""
    labels_col_a = []
    labels_col_b = []
    date_cells = {}  # (row,col) -> iso
    for r in range(1, min(max_rows, ws.max_row) + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        if a is not None:
            labels_col_a.append((r, safe_str(a)[:60]))
        if b is not None:
            labels_col_b.append((r, safe_str(b)[:60]))
        for c in range(1, min(max_cols, ws.max_column) + 1):
            v = ws.cell(r, c).value
            if isinstance(v, (datetime, date)):
                date_cells[f"R{r}C{c}"] = safe_str(v)
    return {
        "dimensions": f"{ws.max_row}x{ws.max_column}",
        "labels_col_a": labels_col_a[:40],
        "labels_col_b": labels_col_b[:40],
        "date_cells": date_cells,
    }

def profile_file(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    result = {
        "sheetnames": list(wb.sheetnames),
        "sheet_count": len(wb.sheetnames),
        "core_sheets": {},
    }
    for core in CORE_SHEETS:
        actual = sheet_lookup(wb, core)
        if actual is None:
            result["core_sheets"][core] = {"present": False}
            continue
        ws = wb[actual]
        profile = profile_sheet(ws)
        profile["present"] = True
        profile["actual_name"] = actual
        profile["trailing_space"] = actual != actual.strip()
        result["core_sheets"][core] = profile
    wb.close()
    return result

# -------------------------------------------------------------------
# PROFILE ALL FILES
# -------------------------------------------------------------------
print(f"Profiling {len(FILES)} files...")
profiles = OrderedDict()
for period, fname in FILES:
    path = os.path.join(UPLOADS_DIR, fname)
    if not os.path.exists(path):
        print(f"  [MISSING] {period}  {fname}")
        profiles[period] = {"error": "FILE_NOT_FOUND", "filename": fname}
        continue
    print(f"  [OK] {period}  {fname}")
    try:
        profiles[period] = profile_file(path)
        profiles[period]["filename"] = fname
    except Exception as e:
        profiles[period] = {"error": str(e), "filename": fname}

# -------------------------------------------------------------------
# BUILD COMPARISON REPORT (Feb 2026 = canonical)
# -------------------------------------------------------------------
canonical = profiles[CANONICAL_KEY]
canonical_sheets = set(canonical["sheetnames"])
canonical_core = canonical["core_sheets"]

# Canonical schema: expose labels positions for each core sheet
schema = {"canonical_period": CANONICAL_KEY, "canonical_file": canonical["filename"]}
schema["core_sheets"] = {}
for core, data in canonical_core.items():
    if not data.get("present"):
        continue
    schema["core_sheets"][core] = {
        "actual_name": data["actual_name"],
        "trailing_space": data["trailing_space"],
        "dimensions": data["dimensions"],
        "labels_col_a": data["labels_col_a"],
        "labels_col_b": data["labels_col_b"],
        "date_cells": data["date_cells"],
    }

with open(os.path.join(OUT_DIR, "canonical_schema.json"), "w") as f:
    json.dump(schema, f, indent=2, default=str)
print(f"\nCanonical schema -> {OUT_DIR}/canonical_schema.json")

# -------------------------------------------------------------------
# COMPARISON REPORT (markdown)
# -------------------------------------------------------------------
lines = []
lines.append(f"# MA File Profile Report")
lines.append(f"_Canonical reference: **{canonical['filename']}** ({CANONICAL_KEY})_\n")

# 1. Sheet-presence matrix
lines.append("## 1. Core sheet presence per file\n")
header = "| Period | " + " | ".join(CORE_SHEETS) + " | Total sheets |"
sep = "|" + "---|" * (len(CORE_SHEETS) + 2)
lines.append(header)
lines.append(sep)
for period, prof in profiles.items():
    if "error" in prof:
        row = f"| **{period}** | ERROR: {prof['error']} |"
        lines.append(row)
        continue
    cells = []
    for core in CORE_SHEETS:
        d = prof["core_sheets"].get(core, {})
        if not d.get("present"):
            cells.append("MISSING")
        else:
            mark = "OK"
            if d.get("trailing_space"):
                mark = f"OK *(name: `{d['actual_name']}`)*"
            elif d["actual_name"] != core:
                mark = f"`{d['actual_name']}`"
            cells.append(mark)
    row = f"| **{period}** | " + " | ".join(cells) + f" | {prof['sheet_count']} |"
    lines.append(row)

# 2. Sheet-name drift (any sheet names across all files that differ from canonical only by whitespace/case)
lines.append("\n## 2. Sheet-name drift warnings\n")
for period, prof in profiles.items():
    if "error" in prof:
        continue
    drift = []
    for s in prof["sheetnames"]:
        if s not in canonical_sheets and s.strip() in canonical_sheets:
            drift.append(f"`{s}` (trailing/leading whitespace)")
        elif s not in canonical_sheets and s.lower().strip() in {c.lower().strip() for c in canonical_sheets}:
            drift.append(f"`{s}` (case drift)")
    if drift:
        lines.append(f"- **{period}**: " + ", ".join(drift))

# 3. Date-cell locations per core sheet
lines.append("\n## 3. Date cell location per file (core sheets)\n")
for core in CORE_SHEETS:
    lines.append(f"\n### {core}\n")
    lines.append("| Period | Date cells found |")
    lines.append("|---|---|")
    for period, prof in profiles.items():
        if "error" in prof:
            continue
        d = prof["core_sheets"].get(core, {})
        if not d.get("present"):
            lines.append(f"| {period} | _MISSING_ |")
            continue
        dc = d.get("date_cells", {})
        if not dc:
            lines.append(f"| {period} | _no date cells_ |")
        else:
            lines.append(f"| {period} | " + ", ".join(f"{k}={v}" for k, v in dc.items()) + " |")

# 4. Label differences vs canonical at key P&L rows
lines.append("\n## 4. P&L Detail — labels in column A per file\n")
canon_pnl = canonical_core.get("P&L Detail", {})
canon_labels_a = dict(canon_pnl.get("labels_col_a", []))
# pick rows we typically care about: 4..30
rows_of_interest = sorted([r for r in canon_labels_a.keys() if 4 <= r <= 30])[:20]
header = "| Period | " + " | ".join(f"R{r}" for r in rows_of_interest) + " |"
sep = "|" + "---|" * (len(rows_of_interest) + 1)
lines.append(header)
lines.append(sep)
# show canonical first
canon_row = "| **CANONICAL** | " + " | ".join(canon_labels_a.get(r, "")[:20] for r in rows_of_interest) + " |"
lines.append(canon_row)
for period, prof in profiles.items():
    if "error" in prof:
        continue
    d = prof["core_sheets"].get("P&L Detail", {})
    if not d.get("present"):
        lines.append(f"| {period} | _MISSING_ |")
        continue
    labels = dict(d.get("labels_col_a", []))
    cells = []
    for r in rows_of_interest:
        v = labels.get(r, "")
        cv = canon_labels_a.get(r, "")
        if v == cv:
            cells.append("=")
        else:
            cells.append(v[:20] if v else "_blank_")
    lines.append(f"| {period} | " + " | ".join(cells) + " |")

with open(os.path.join(OUT_DIR, "profile_report.md"), "w") as f:
    f.write("\n".join(lines))
print(f"Markdown report -> {OUT_DIR}/profile_report.md")

# -------------------------------------------------------------------
# CSV detail dump
# -------------------------------------------------------------------
with open(os.path.join(OUT_DIR, "profile_detail.csv"), "w", newline="") as f:
    w = csv.writer(f)
    w.writerow(["period", "core_sheet", "present", "actual_name", "trailing_space", "dimensions", "date_cells"])
    for period, prof in profiles.items():
        if "error" in prof:
            w.writerow([period, "ERROR", prof["error"], "", "", "", ""])
            continue
        for core in CORE_SHEETS:
            d = prof["core_sheets"].get(core, {})
            w.writerow([
                period, core,
                d.get("present", False),
                d.get("actual_name", ""),
                d.get("trailing_space", ""),
                d.get("dimensions", ""),
                json.dumps(d.get("date_cells", {})),
            ])
print(f"CSV detail      -> {OUT_DIR}/profile_detail.csv")

# -------------------------------------------------------------------
# PRINT TOP-LINE SUMMARY TO STDOUT
# -------------------------------------------------------------------
print("\n" + "=" * 70)
print("TOP-LINE SUMMARY")
print("=" * 70)
for period, prof in profiles.items():
    if "error" in prof:
        print(f"{period}: ERROR {prof['error']}")
        continue
    missing = [c for c in CORE_SHEETS if not prof["core_sheets"].get(c, {}).get("present")]
    drift = [c for c in CORE_SHEETS if prof["core_sheets"].get(c, {}).get("trailing_space")]
    note = ""
    if missing:
        note += f" MISSING={missing}"
    if drift:
        note += f" TRAILING_SPACE={drift}"
    print(f"{period}: {prof['sheet_count']} sheets{note}")
