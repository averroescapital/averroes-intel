import io
import openpyxl
import pandas as pd
from datetime import datetime
import re


def _find_sheet(wb, target, strict=False):
    """
    Find a sheet name tolerantly: case-insensitive, strip trailing whitespace.
    Returns the actual sheet name as it appears in the workbook, or None.
    """
    target_clean = target.strip().lower()
    for name in wb.sheetnames:
        if name.strip().lower() == target_clean:
            return name
    return None


def parse_date(val):
    if isinstance(val, datetime):
        return val.replace(day=1).date().isoformat()
    return "2026-02-01"  # Specific fallback for Portco Alpha MA Feb


def parse_alpha_ma(file_content, file_name):
    """
    Advanced Parser for Portco Alpha (Hospitality SaaS).
    Extracts segmented ARR, CARR, Tech GM%, NRR/GRR, NWC details,
    and per-BL customer/property counts from the Customer Numbers sheet.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    rows_to_insert = []
    reporting_period = None  # resolved from the first available sheet

    # --- 1. SEGMENTED REVENUE (P&L Detail) ---
    pnl_detail = _find_sheet(wb, 'P&L Detail')
    if pnl_detail:
        ws = wb[pnl_detail]
        reporting_period = parse_date(ws.cell(row=3, column=2).value)

        mapping = {
            ' Ecommerce Revenue': 'tech_mrr',
            ' EMS Revenue': 'services_mrr',
            ' Services Revenue': 'other_services_mrr',
            'TOTAL REVENUE': 'total_revenue',
            'Gross Profit': 'total_gross_profit',
            'EBITDA': 'ebitda'
        }

        for r in range(4, 100):
            label = str(ws.cell(row=r, column=1).value).strip()
            if label in mapping:
                val = ws.cell(row=r, column=2).value  # Actual
                if isinstance(val, (int, float)):
                    rows_to_insert.append({
                        "period": reporting_period,
                        "kpi": mapping[label],
                        "value": float(val),
                        "sheet": "P&L Detail"
                    })

    # --- 2. TECH SPECIFIC MARGIN (Tech P&L) ---
    ecom_pnl = _find_sheet(wb, 'Ecommerce P&L')
    if ecom_pnl:
        ws = wb[ecom_pnl]
        if reporting_period is None:
            reporting_period = parse_date(ws.cell(row=3, column=2).value)
        for r in range(4, 50):
            label = str(ws.cell(row=r, column=1).value).strip().lower()
            if 'gross margin' in label:
                val = ws.cell(row=r, column=2).value
                if isinstance(val, (int, float)):
                    rows_to_insert.append({
                        "period": reporting_period,
                        "kpi": "tech_gross_margin_pct",
                        "value": float(val) * 100 if val <= 1 else float(val),
                        "sheet": "Ecommerce P&L"
                    })

    # --- 3. RETENTION & HEALTH (Financial KPIs) ---
    fin_kpis = _find_sheet(wb, 'Financial KPIs')
    if fin_kpis:
        ws = wb[fin_kpis]
        if reporting_period is None:
            reporting_period = parse_date(ws.cell(row=1, column=4).value)
        kpi_map = [
            (3, 4, "carr"),                       # Contracted ARR
            (4, 4, "live_arr"),                   # Billing ARR
            (6, 4, "nrr_pct"),                    # Net Retention
            (7, 4, "grr_pct"),                    # Gross Retention
            (10, 4, "customer_concentration_pct"), # Top 10 focus
            (12, 4, "nps"),                        # Customer satisfaction
        ]
        for r, c, kpi in kpi_map:
            val = ws.cell(row=r, column=c).value
            if isinstance(val, (int, float)):
                rows_to_insert.append({
                    "period": reporting_period,
                    "kpi": kpi,
                    "value": float(val),
                    "sheet": "Financial KPIs"
                })

    # --- 4. MODULES & LTV:CAC (Hospitality Metrics) ---
    hosp_metrics = _find_sheet(wb, 'Hospitality Metrics')
    if hosp_metrics:
        ws = wb[hosp_metrics]
        if reporting_period is None:
            reporting_period = parse_date(ws.cell(row=3, column=2).value)
        h_map = [
            (5, 2, "rooms_module_delta"),
            (6, 2, "spa_module_delta"),
            (7, 2, "f_b_module_delta"),
            (8, 2, "vouchers_module_delta"),
            (15, 2, "ltv_cac_ratio"),
            (16, 2, "cac_payback_months")
        ]
        for r, c, kpi in h_map:
            val = ws.cell(row=r, column=c).value
            if isinstance(val, (int, float)):
                rows_to_insert.append({
                    "period": reporting_period,
                    "kpi": kpi,
                    "value": float(val),
                    "sheet": "Hospitality Metrics"
                })

    # --- 5. NWC & DEBTOR AGING (Balance Sheet) ---
    bs = _find_sheet(wb, 'Balance Sheet')
    if bs:
        ws = wb[bs]
        if reporting_period is None:
            reporting_period = parse_date(ws.cell(row=3, column=3).value)
        for r in range(4, 60):
            label = str(ws.cell(row=r, column=1).value).strip().lower()
            if 'debtors' in label:
                rows_to_insert.append({"period": reporting_period, "kpi": "debtors_30d", "value": float(ws.cell(row=r, column=4).value or 0), "sheet": "Balance Sheet"})
                rows_to_insert.append({"period": reporting_period, "kpi": "debtors_60d", "value": float(ws.cell(row=r, column=5).value or 0), "sheet": "Balance Sheet"})
                rows_to_insert.append({"period": reporting_period, "kpi": "debtors_90d", "value": float(ws.cell(row=r, column=6).value or 0), "sheet": "Balance Sheet"})

    # --- 6. CUSTOMER NUMBERS (per-BL property/module counts) ---
    # Cross-tab sheet: each column is a month, rows hold BL-level counts.
    # We extract ALL populated months (not just the current period) so that
    # the backfill run surfaces historical customer numbers from Era 3 files.
    # Dedup is handled downstream in silver (latest ingested_at wins).
    #
    # Layout (Feb 2026 canonical):
    #   R2: end-of-month dates  |  R3: first-of-month dates  |  cols B(2)..Y(25)
    #   R4: "ACTUAL PROPERTIES"
    #     R5: Ecom  R6: EMS  R7: Services  R8: Total
    #   R10: "ACTUAL REVENUE" (absolute £, not £k)
    #     R11: Ecom  R12: EMS  R13: Services  R14: Total
    #   R16: "AVE REVENUE PER CUSTOMER"
    #     R17: Ecom  R18: EMS  R19: Services  R20: Total
    #   R22: "BUDGET PROPERTIES"
    #     R23: Ecom  R24: EMS  R25: Services  R26: Total
    #   R28+: Geo breakdowns (UK, Ireland, Italy, Spain/UAE)
    cust_sheet = _find_sheet(wb, 'Customer Numbers')
    if cust_sheet:
        ws = wb[cust_sheet]

        # Build column→period map from R3 (first-of-month dates)
        col_periods = {}
        for c in range(2, ws.max_column + 1):
            date_val = ws.cell(row=3, column=c).value
            if isinstance(date_val, datetime):
                col_periods[c] = date_val.replace(day=1).date().isoformat()
            # Skip YTD column (col 26 typically) and any non-date

        # ---- ACTUAL PROPERTIES → modules_live_* ----
        actual_props = {
            5: ("modules_live_ecommerce", "ecommerce"),
            6: ("modules_live_ems",       "ems"),
            7: ("modules_live_services",  "services"),
            8: ("modules_live_total",     "total"),
        }
        for row_num, (kpi_name, bl) in actual_props.items():
            for c, period_str in col_periods.items():
                val = ws.cell(row=row_num, column=c).value
                if isinstance(val, (int, float)):
                    rows_to_insert.append({
                        "period": period_str,
                        "kpi": kpi_name,
                        "value": float(val),
                        "sheet": "Customer Numbers",
                        "source_cell": ws.cell(row=row_num, column=c).coordinate,
                        "business_line": bl,
                    })

        # ---- BUDGET PROPERTIES → modules_budget_* ----
        budget_props = {
            23: ("modules_budget_ecommerce", "ecommerce"),
            24: ("modules_budget_ems",       "ems"),
            25: ("modules_budget_services",  "services"),
            26: ("modules_budget_total",     "total"),
        }
        for row_num, (kpi_name, bl) in budget_props.items():
            for c, period_str in col_periods.items():
                val = ws.cell(row=row_num, column=c).value
                if isinstance(val, (int, float)) and val != 0:
                    rows_to_insert.append({
                        "period": period_str,
                        "kpi": kpi_name,
                        "value": float(val),
                        "sheet": "Customer Numbers",
                        "source_cell": ws.cell(row=row_num, column=c).coordinate,
                        "business_line": bl,
                        "value_type": "budget",
                    })

        # ---- ACTUAL REVENUE per BL (absolute £, convert to £k) ----
        actual_rev = {
            11: ("customer_revenue_ecommerce", "ecommerce"),
            12: ("customer_revenue_ems",       "ems"),
            13: ("customer_revenue_services",  "services"),
            14: ("customer_revenue_total",     "total"),
        }
        for row_num, (kpi_name, bl) in actual_rev.items():
            for c, period_str in col_periods.items():
                val = ws.cell(row=row_num, column=c).value
                if isinstance(val, (int, float)):
                    rows_to_insert.append({
                        "period": period_str,
                        "kpi": kpi_name,
                        "value": round(float(val) / 1000.0, 5),  # £ → £k
                        "sheet": "Customer Numbers",
                        "source_cell": ws.cell(row=row_num, column=c).coordinate,
                        "business_line": bl,
                    })

        # ---- AVE REVENUE PER CUSTOMER per BL ----
        arpc = {
            17: ("arpc_ecommerce", "ecommerce"),
            18: ("arpc_ems",       "ems"),
            19: ("arpc_services",  "services"),
            20: ("arpc_total",     "total"),
        }
        for row_num, (kpi_name, bl) in arpc.items():
            for c, period_str in col_periods.items():
                val = ws.cell(row=row_num, column=c).value
                if isinstance(val, (int, float)) and val != 0:
                    rows_to_insert.append({
                        "period": period_str,
                        "kpi": kpi_name,
                        "value": float(val),
                        "sheet": "Customer Numbers",
                        "source_cell": ws.cell(row=row_num, column=c).coordinate,
                        "business_line": bl,
                    })

        # ---- GEO BREAKDOWN: actual property counts per region ----
        geo_blocks = {
            "uk":          {"ecom": 29, "ems": 30, "services": 31, "total": 32},
            "ireland":     {"ecom": 35, "ems": 36, "services": 37, "total": 38},
            "italy":       {"ecom": 41, "ems": 42, "services": 43, "total": 44},
            "spain_uae":   {"ecom": 47, "ems": 48, "services": 49, "total": 50},
        }
        for geo, bl_rows in geo_blocks.items():
            for bl_key, row_num in bl_rows.items():
                kpi_name = f"properties_{geo}_{bl_key}"
                for c, period_str in col_periods.items():
                    val = ws.cell(row=row_num, column=c).value
                    if isinstance(val, (int, float)) and val != 0:
                        rows_to_insert.append({
                            "period": period_str,
                            "kpi": kpi_name,
                            "value": float(val),
                            "sheet": "Customer Numbers",
                            "source_cell": ws.cell(row=row_num, column=c).coordinate,
                            "business_line": bl_key if bl_key != "total" else "total",
                        })

    # --- 7. GL COVENANTS ---
    # Layout: Labels in cols B and C. Section headers in B ('ARR', 'Interest Cover',
    # 'Debt Service Ratio Cover', 'Cash Minimum Balance'). Within ARR section,
    # total row has NO labels (A/B/C empty) with D=actual, E=covenant.
    gl_cov = _find_sheet(wb, 'GL Covenants')
    if gl_cov:
        ws = wb[gl_cov]
        if reporting_period is None:
            reporting_period = parse_date(ws.cell(row=3, column=2).value)

        section = None
        arr_total_found = False
        for r in range(1, min(60, ws.max_row) + 1):
            lab_a = str(ws.cell(row=r, column=1).value or '').strip().lower()
            lab_b = str(ws.cell(row=r, column=2).value or '').strip().lower()
            lab_c = str(ws.cell(row=r, column=3).value or '').strip().lower()
            combined = f"{lab_a} {lab_b} {lab_c}"

            # Section detection
            if lab_b == 'arr' or 'arr covenant' in combined:
                section = 'arr'
                arr_total_found = False
                continue
            elif 'interest cover' in lab_b and 'debt' not in combined:
                section = 'interest'
                continue
            elif 'debt service' in lab_b or 'debt service' in combined:
                section = 'debt'
                continue
            elif 'cash minimum' in lab_b or 'cash min' in combined:
                section = 'cash_min'
                continue

            if section == 'arr':
                if not arr_total_found:
                    val = ws.cell(row=r, column=4).value
                    val2 = ws.cell(row=r, column=5).value
                    has_label = bool(lab_a or lab_b or lab_c)
                    if not has_label and isinstance(val, (int, float)) and isinstance(val2, (int, float)) and val > 1000:
                        rows_to_insert.append({"period": reporting_period, "kpi": "gl_arr_actual", "value": float(val), "sheet": "GL Covenants"})
                        rows_to_insert.append({"period": reporting_period, "kpi": "gl_arr_covenant", "value": float(val2), "sheet": "GL Covenants"})
                        arr_total_found = True
                        continue
                if 'covenant' in lab_c:
                    val = ws.cell(row=r, column=4).value
                    if isinstance(val, (int, float)) and val < 2:
                        rows_to_insert.append({"period": reporting_period, "kpi": "gl_arr_threshold", "value": float(val), "sheet": "GL Covenants"})
                elif 'actual' in lab_c:
                    val = ws.cell(row=r, column=4).value
                    if isinstance(val, (int, float)) and val < 2:
                        rows_to_insert.append({"period": reporting_period, "kpi": "gl_arr_ratio", "value": float(val), "sheet": "GL Covenants"})

            elif section == 'interest':
                if 'interest' in lab_c and ('charge' in lab_c or 'expense' in lab_c or 'cost' in lab_c):
                    val = ws.cell(row=r, column=4).value
                    if isinstance(val, (int, float)):
                        rows_to_insert.append({"period": reporting_period, "kpi": "gl_interest_cover_interest", "value": float(val), "sheet": "GL Covenants"})
                elif 'ebitda' in lab_c:
                    val = ws.cell(row=r, column=4).value
                    if isinstance(val, (int, float)):
                        rows_to_insert.append({"period": reporting_period, "kpi": "gl_interest_cover_ebitda", "value": float(val), "sheet": "GL Covenants"})
                elif 'interest cover' in lab_c:
                    val = ws.cell(row=r, column=4).value
                    if isinstance(val, (int, float)) and abs(val) < 100:
                        rows_to_insert.append({"period": reporting_period, "kpi": "gl_interest_cover_ratio", "value": float(val), "sheet": "GL Covenants"})

            elif section == 'debt':
                if 'debt service ratio' in lab_c or 'ratio' in lab_c:
                    val = ws.cell(row=r, column=4).value
                    if isinstance(val, (int, float)):
                        rows_to_insert.append({"period": reporting_period, "kpi": "gl_debt_service_ratio", "value": float(val), "sheet": "GL Covenants"})

            elif section == 'cash_min':
                if 'covenant' in lab_c:
                    val = ws.cell(row=r, column=4).value
                    if isinstance(val, (int, float)):
                        rows_to_insert.append({"period": reporting_period, "kpi": "gl_cash_min_balance", "value": float(val), "sheet": "GL Covenants"})

    # --- 8. AVERROES GUARD RAILS (full covenant compliance) ---
    # Layout: Section headers in col B ('Revenue','MRR','Contribution',
    # 'EBITDA less Capex','Cash Balance'). Within each: Covenant/Actual/Ratio
    # in col D. Stop at 'KPIs' section (different layout).
    gr_sheet = _find_sheet(wb, 'Averroes Guard Rails')
    if gr_sheet:
        ws = wb[gr_sheet]
        if reporting_period is None:
            reporting_period = parse_date(ws.cell(row=3, column=2).value)

        gr_blocks = [
            ('revenue',      'gr_revenue_covenant_ytd',      'gr_revenue_actual_ytd',      'gr_revenue_ratio'),
            ('mrr',          'gr_mrr_covenant',              'gr_mrr_actual',              'gr_mrr_ratio'),
            ('contribution', 'gr_contribution_covenant_ytd', 'gr_contribution_actual_ytd', 'gr_contribution_ratio'),
            ('ebitda',       'gr_ebitda_capex_covenant_ytd', 'gr_ebitda_capex_actual_ytd', 'gr_ebitda_capex_ratio'),
            ('cash',         'gr_cash_covenant',             'gr_cash_actual',             'gr_cash_ratio'),
        ]

        current_block = None
        current_kpis = None
        block_idx = 0

        for r in range(1, min(55, ws.max_row) + 1):
            lab_b = str(ws.cell(row=r, column=2).value or '').strip().lower()
            lab_c = str(ws.cell(row=r, column=3).value or '').strip().lower()

            # Detect section headers (col B, no col C label)
            if lab_b and not lab_c:
                if lab_b == 'kpis':
                    break  # Stop — KPIs section uses different layout
                for keyword, *kpis in gr_blocks:
                    if keyword in lab_b:
                        current_block = keyword
                        current_kpis = kpis
                        block_idx = 0
                        break
                continue

            if current_block is None or current_kpis is None:
                continue

            val = ws.cell(row=r, column=4).value
            if isinstance(val, (int, float)) and block_idx < 3:
                rows_to_insert.append({
                    "period": reporting_period,
                    "kpi": current_kpis[block_idx],
                    "value": float(val),
                    "sheet": "Averroes Guard Rails",
                    "source_cell": ws.cell(row=r, column=4).coordinate,
                })
                block_idx += 1
                if block_idx >= 3:
                    current_block = None

    return rows_to_insert
