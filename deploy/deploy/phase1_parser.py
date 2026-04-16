"""
Phase 1 MA Parser — Extracts KPIs from MAfileFeb26.xlsx format
into gold.kpi_monthly schema (Phase 1 KPI Dictionary: Sections A-E).

Input:  MAfile<Mon><YY>.xlsx (36 sheets, monthly management accounts)
Output: Dict matching gold.kpi_monthly columns

FY = Nov-Oct. Q1=Nov-Jan, Q2=Feb-Apr, Q3=May-Jul, Q4=Aug-Oct
Currency: GBP, £k (P&L) or £ (Financial KPIs, Cash, NWC)
"""

import io
import openpyxl
import pandas as pd
from datetime import datetime


def get_fy_info(period_date):
    """Derive FY, quarter, month number from a date. FY starts Nov."""
    m = period_date.month
    y = period_date.year
    if m >= 11:
        fy = f"FY{(y + 1) % 100:02d}"
        fy_month = m - 10  # Nov=1, Dec=2
    else:
        fy = f"FY{y % 100:02d}"
        fy_month = m + 2   # Jan=3, Feb=4, ... Oct=12

    if fy_month <= 3:
        q = "Q1"
    elif fy_month <= 6:
        q = "Q2"
    elif fy_month <= 9:
        q = "Q3"
    else:
        q = "Q4"

    return fy, q, fy_month


def safe_float(val):
    """Convert cell value to float, return None if not numeric."""
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def parse_ma_file(file_path_or_bytes, portco_id="portco-alpha"):
    """
    Parse a Management Accounts Excel file and return a dict
    matching gold.kpi_monthly columns.

    Args:
        file_path_or_bytes: Path string or bytes of the Excel file
        portco_id: Portfolio company identifier

    Returns:
        dict with all KPI values for the reporting period
    """
    if isinstance(file_path_or_bytes, bytes):
        wb = openpyxl.load_workbook(io.BytesIO(file_path_or_bytes), data_only=True)
    else:
        wb = openpyxl.load_workbook(file_path_or_bytes, data_only=True)

    row = {"portco_id": portco_id}

    # ─────────────────────────────────────────────
    # 1. REPORTING PERIOD (from P&L Summary, cell B3)
    # ─────────────────────────────────────────────
    ws_pnl = wb["P&L Summary "]
    period_raw = ws_pnl.cell(row=3, column=2).value
    if isinstance(period_raw, datetime):
        period = period_raw.date()
    else:
        period = pd.to_datetime(str(period_raw)).date()

    row["period"] = str(period)
    fy, q, fy_month = get_fy_info(period)
    row["fy"] = fy
    row["fy_quarter"] = q
    row["fy_month_num"] = fy_month

    # ─────────────────────────────────────────────
    # 2. SECTION A: ARR / MRR / SCALE
    # ─────────────────────────────────────────────
    ws_kpi = wb["Financial KPIs"]

    # Block 1 (row 3-6): TECH MRR / SERVICES MRR
    # Tech MRR - Current Month: C3=header, row4=prior, row5=budget, row6=actual
    row["tech_mrr_prior_year"] = safe_float(ws_kpi.cell(row=4, column=3).value)
    row["tech_mrr_budget"] = safe_float(ws_kpi.cell(row=5, column=3).value)
    row["tech_mrr_actual"] = safe_float(ws_kpi.cell(row=6, column=3).value)

    # Tech MRR - YTD (cols F/G)
    row["tech_mrr_ytd_prior_year"] = safe_float(ws_kpi.cell(row=4, column=7).value)
    row["tech_mrr_ytd_budget"] = safe_float(ws_kpi.cell(row=5, column=7).value)
    row["tech_mrr_ytd_actual"] = safe_float(ws_kpi.cell(row=6, column=7).value)

    # Services MRR - Current Month (cols J/K)
    row["services_mrr_prior_year"] = safe_float(ws_kpi.cell(row=4, column=11).value)
    row["services_mrr_budget"] = safe_float(ws_kpi.cell(row=5, column=11).value)
    row["services_mrr_actual"] = safe_float(ws_kpi.cell(row=6, column=11).value)

    # Services MRR - YTD (cols N/O)
    row["services_mrr_ytd_prior_year"] = safe_float(ws_kpi.cell(row=4, column=15).value)
    row["services_mrr_ytd_budget"] = safe_float(ws_kpi.cell(row=5, column=15).value)
    row["services_mrr_ytd_actual"] = safe_float(ws_kpi.cell(row=6, column=15).value)

    # Derived: Total MRR and YTD-based ARR
    t_mrr_actual = row["tech_mrr_actual"] or 0
    s_mrr_actual = row["services_mrr_actual"] or 0
    row["total_mrr_actual"] = t_mrr_actual + s_mrr_actual

    # ARR now uses YTD values directly as per request
    t_mrr_ytd = row["tech_mrr_ytd_actual"] or 0
    s_mrr_ytd = row["services_mrr_ytd_actual"] or 0
    row["tech_arr"] = t_mrr_ytd
    row["services_arr"] = s_mrr_ytd
    row["total_arr"] = t_mrr_ytd + s_mrr_ytd

    # Revenue from P&L Summary (£k)
    # Col B=Actual, C=Budget, E=Prior Year (month)
    # Col G=YTD Actual, H=YTD Budget, J=YTD Prior Year
    row["revenue_ecommerce_actual"] = safe_float(ws_pnl.cell(row=5, column=2).value)
    row["revenue_ecommerce_budget"] = safe_float(ws_pnl.cell(row=5, column=3).value)
    row["revenue_ems_actual"] = safe_float(ws_pnl.cell(row=6, column=2).value)
    row["revenue_ems_budget"] = safe_float(ws_pnl.cell(row=6, column=3).value)
    row["revenue_services_actual"] = safe_float(ws_pnl.cell(row=7, column=2).value)
    row["revenue_services_budget"] = safe_float(ws_pnl.cell(row=7, column=3).value)
    row["revenue_total_actual"] = safe_float(ws_pnl.cell(row=8, column=2).value)
    row["revenue_total_budget"] = safe_float(ws_pnl.cell(row=8, column=3).value)
    row["revenue_total_prior_year"] = safe_float(ws_pnl.cell(row=8, column=5).value)

    # YTD Revenue
    row["revenue_total_ytd_actual"] = safe_float(ws_pnl.cell(row=8, column=7).value)
    row["revenue_total_ytd_budget"] = safe_float(ws_pnl.cell(row=8, column=8).value)
    row["revenue_total_ytd_prior_year"] = safe_float(ws_pnl.cell(row=8, column=10).value)

    # Revenue growth %
    rev_act = row["revenue_total_actual"] or 0
    rev_bud = row["revenue_total_budget"] or 0
    rev_py = row["revenue_total_prior_year"] or 0
    row["revenue_vs_budget_pct"] = ((rev_act / rev_bud) - 1) * 100 if rev_bud else None
    row["revenue_yoy_growth_pct"] = ((rev_act / rev_py) - 1) * 100 if rev_py else None

    ytd_act = row["revenue_total_ytd_actual"] or 0
    ytd_py = row["revenue_total_ytd_prior_year"] or 0
    row["revenue_ytd_growth_pct"] = ((ytd_act / ytd_py) - 1) * 100 if ytd_py else None

    # ─────────────────────────────────────────────
    # 3. SECTION B: UNIT ECONOMICS / MARGINS
    # ─────────────────────────────────────────────
    # Block 2 (row 19-22): AVE REVENUE PER CUSTOMER / YTD REVENUE GROWTH / S&M Efficiency
    row["arpc_budget"] = safe_float(ws_kpi.cell(row=21, column=3).value)
    row["arpc_actual"] = safe_float(ws_kpi.cell(row=22, column=3).value)
    row["arpc_ytd_budget"] = safe_float(ws_kpi.cell(row=21, column=7).value)
    row["arpc_ytd_actual"] = safe_float(ws_kpi.cell(row=22, column=7).value)

    # YTD Revenue Growth (col K) — already captured above
    # S&M Efficiency (col O)
    row["sm_efficiency"] = safe_float(ws_kpi.cell(row=22, column=15).value)
    row["sm_efficiency_ytd"] = safe_float(ws_kpi.cell(row=22, column=15).value)  # same cell for now

    # Block 3 (row 35-38): TECH GROSS MARGIN / EBITDA MARGIN
    row["tech_gross_margin_prior_pct"] = safe_float(ws_kpi.cell(row=36, column=3).value)
    row["tech_gross_margin_budget_pct"] = safe_float(ws_kpi.cell(row=37, column=3).value)
    row["tech_gross_margin_pct"] = safe_float(ws_kpi.cell(row=38, column=3).value)

    row["tech_gross_margin_ytd_pct"] = safe_float(ws_kpi.cell(row=38, column=7).value)

    row["ebitda_margin_prior_pct"] = safe_float(ws_kpi.cell(row=36, column=11).value)
    row["ebitda_margin_budget_pct"] = safe_float(ws_kpi.cell(row=37, column=11).value)
    row["ebitda_margin_pct"] = safe_float(ws_kpi.cell(row=38, column=11).value)
    row["ebitda_margin_ytd_pct"] = safe_float(ws_kpi.cell(row=38, column=15).value)

    # P&L: Direct costs, Gross Profit, Contribution, EBITDA
    row["direct_costs_total"] = safe_float(ws_pnl.cell(row=12, column=2).value)
    row["gross_profit_ecommerce"] = safe_float(ws_pnl.cell(row=13, column=2).value)
    row["gross_profit_ems"] = safe_float(ws_pnl.cell(row=14, column=2).value)
    row["gross_profit_services"] = safe_float(ws_pnl.cell(row=15, column=2).value)

    gp_ecom = row["gross_profit_ecommerce"] or 0
    gp_ems = row["gross_profit_ems"] or 0
    gp_svc = row["gross_profit_services"] or 0
    row["gross_profit_total"] = gp_ecom + gp_ems + gp_svc

    row["contribution_ecommerce"] = safe_float(ws_pnl.cell(row=13, column=2).value)
    row["contribution_ems"] = safe_float(ws_pnl.cell(row=14, column=2).value)
    row["contribution_services"] = safe_float(ws_pnl.cell(row=15, column=2).value)
    row["contribution_total"] = safe_float(ws_pnl.cell(row=16, column=2).value)

    cont_total = row["contribution_total"] or 0
    row["contribution_margin_pct"] = (cont_total / rev_act) * 100 if rev_act else None

    row["total_overheads"] = safe_float(ws_pnl.cell(row=17, column=2).value)
    row["ebitda_actual"] = safe_float(ws_pnl.cell(row=18, column=2).value)
    row["ebitda_budget"] = safe_float(ws_pnl.cell(row=18, column=3).value)
    row["ebitda_prior_year"] = safe_float(ws_pnl.cell(row=18, column=5).value)
    row["capex"] = safe_float(ws_pnl.cell(row=19, column=2).value)
    row["ebitda_less_capex"] = safe_float(ws_pnl.cell(row=20, column=2).value)

    # EBITDA YTD
    row["ebitda_ytd_actual"] = safe_float(ws_pnl.cell(row=18, column=7).value)
    row["ebitda_ytd_budget"] = safe_float(ws_pnl.cell(row=18, column=8).value)

    # ─────────────────────────────────────────────
    # 4. SECTION C: RETENTION / CHURN
    # ─────────────────────────────────────────────
    # Block 5 (row 67-71): INDICATIVE EV / RULE OF 40 / REVENUE CHURN / TIME TO VALUE
    row["revenue_churn_target"] = safe_float(ws_kpi.cell(row=69, column=11).value)
    row["revenue_churn_pct"] = safe_float(ws_kpi.cell(row=70, column=11).value)

    # Rule of 40
    arr_growth = safe_float(ws_kpi.cell(row=68, column=7).value)
    ebitda_margin_r40 = safe_float(ws_kpi.cell(row=69, column=7).value)
    rule_40_val = safe_float(ws_kpi.cell(row=70, column=7).value)
    row["rule_of_40"] = rule_40_val

    # Indicative EV
    row["indicative_ev"] = safe_float(ws_kpi.cell(row=69, column=3).value)

    # Time to Value
    row["time_to_value_days"] = safe_float(ws_kpi.cell(row=70, column=15).value)
    row["time_to_value_excl_blocked"] = safe_float(ws_kpi.cell(row=71, column=15).value)

    # ─────────────────────────────────────────────
    # 5. SECTION E: CAPITAL EFFICIENCY / CASH / NWC
    # ─────────────────────────────────────────────
    # Block 4 (row 51-54): NWC / CASH / FREE CASH CONVERSION
    row["nwc_prior_month"] = safe_float(ws_kpi.cell(row=52, column=3).value)
    row["nwc_budget"] = safe_float(ws_kpi.cell(row=53, column=3).value)
    row["net_working_capital"] = safe_float(ws_kpi.cell(row=54, column=3).value)

    row["cash_balance_prior_month"] = safe_float(ws_kpi.cell(row=52, column=7).value)
    row["cash_balance_budget"] = safe_float(ws_kpi.cell(row=53, column=7).value)
    row["cash_balance"] = safe_float(ws_kpi.cell(row=54, column=7).value)

    # Cash burn = current - prior month
    cb = row["cash_balance"] or 0
    cb_prev = row["cash_balance_prior_month"] or 0
    row["cash_burn_monthly"] = cb - cb_prev

    # Runway = cash / abs(monthly EBITDA) if negative
    ebitda_val = row["ebitda_actual"] or 0
    if ebitda_val < 0:
        row["cash_runway_months"] = cb / (abs(ebitda_val) * 1000) if ebitda_val != 0 else None  # cash in £, ebitda in £k
    else:
        row["cash_runway_months"] = None  # profitable, no burn

    # Free Cash Conversion: The native Excel formula missed the '000. 
    # We multiply by 1000 to neutralize the unit mismatch.
    fcc_budget = safe_float(ws_kpi.cell(row=53, column=11).value)
    row["free_cash_conversion_budget"] = fcc_budget * 1000.0 if fcc_budget else None
    
    fcc_month = safe_float(ws_kpi.cell(row=54, column=11).value)
    row["free_cash_conversion_month"] = fcc_month * 1000.0 if fcc_month else None
    
    fcc_ytd = safe_float(ws_kpi.cell(row=54, column=15).value)
    row["free_cash_conversion_ytd"] = fcc_ytd * 1000.0 if fcc_ytd else None


    # ─────────────────────────────────────────────
    # 6. HEADCOUNT
    # ─────────────────────────────────────────────
    ws_hc = wb["Headcount"]
    # Row 27 = totals (actual col C, budget col D)
    row["total_headcount"] = safe_float(ws_hc.cell(row=27, column=3).value)
    row["headcount_budget"] = safe_float(ws_hc.cell(row=27, column=4).value)

    # Segment headcount (sum by rows)
    ecom_hc = sum(safe_float(ws_hc.cell(row=r, column=3).value) or 0 for r in range(4, 11))
    ems_hc = sum(safe_float(ws_hc.cell(row=r, column=3).value) or 0 for r in range(11, 17))
    svc_hc = sum(safe_float(ws_hc.cell(row=r, column=3).value) or 0 for r in range(17, 22))
    central_hc = sum(safe_float(ws_hc.cell(row=r, column=3).value) or 0 for r in range(22, 27))

    row["headcount_ecommerce"] = ecom_hc
    row["headcount_ems"] = ems_hc
    row["headcount_services"] = svc_hc
    row["headcount_central"] = central_hc

    # Gross Payroll (row 58)
    row["gross_payroll"] = safe_float(ws_hc.cell(row=58, column=3).value)
    row["gross_payroll_budget"] = safe_float(ws_hc.cell(row=58, column=4).value)

    # Revenue per employee
    row["revenue_per_employee"] = safe_float(ws_hc.cell(row=29, column=3).value)
    row["payroll_pct_revenue"] = safe_float(ws_hc.cell(row=60, column=3).value)

    # ─────────────────────────────────────────────
    # 7. SECTION D: MODULES (from Key Asset Data or Customer Numbers)
    # ─────────────────────────────────────────────
    if "Customer Numbers" in wb.sheetnames:
        ws_cust = wb["Customer Numbers"]
        # Find the column for the reporting period
        # Row 2 has dates, columns B onwards
        target_col = None
        for c in range(2, ws_cust.max_column + 1):
            cell_val = ws_cust.cell(row=2, column=c).value
            if isinstance(cell_val, datetime):
                if cell_val.year == period.year and cell_val.month == period.month:
                    target_col = c
                    break

        if target_col:
            # Properties (rows 5-8) - Use row 8 for deduped total
            row["properties_ecommerce"] = safe_float(ws_cust.cell(row=5, column=target_col).value)
            row["properties_ems"] = safe_float(ws_cust.cell(row=6, column=target_col).value)
            row["properties_services"] = safe_float(ws_cust.cell(row=7, column=target_col).value)
            props_total = safe_float(ws_cust.cell(row=8, column=target_col).value)
            row["properties_live"] = int(props_total) if props_total else None

    # ─────────────────────────────────────────────
    # 8. CARR / Implementation Backlog (placeholder - requires pipeline data)
    # ─────────────────────────────────────────────
    # CARR = ARR + signed-not-live pipeline. Not directly in MA file.
    # Set to ARR for now, update when pipeline data available.
    row["carr"] = row["total_arr"]
    row["implementation_backlog"] = 0

    # ─────────────────────────────────────────────
    # 9. KPI DATA SHEET (S&M, TCV, CAC)
    # ─────────────────────────────────────────────
    if "KPI data" in wb.sheetnames:
        ws_kd = wb["KPI data"]
        # Determine which quarter column to use based on fy_quarter
        # R1: headers = Q1 FY26, Q2 FY26, Q3 FY26, Q4 FY26, YTD
        # Cols: B=Q1, C=Q2, D=Q3, E=Q4, F=YTD
        q_col_map = {"Q1": 2, "Q2": 3, "Q3": 4, "Q4": 5}
        q_col = q_col_map.get(q, 3)  # default Q2 for Feb

        mrr_new = safe_float(ws_kd.cell(row=2, column=q_col).value)
        tcv = safe_float(ws_kd.cell(row=6, column=q_col).value)
        sales_cost = safe_float(ws_kd.cell(row=8, column=q_col).value)
        sm_eff = safe_float(ws_kd.cell(row=10, column=q_col).value)

        if sm_eff:
            row["sm_efficiency"] = sm_eff

        # CAC = Sales Cost / New Customers (approximation)
        if sales_cost and mrr_new and mrr_new > 0:
            # Estimate new customers = MRR_new / ARPC
            arpc = row["arpc_actual"] or 1200
            new_custs = (mrr_new / arpc) if arpc > 0 else 1
            row["cac"] = sales_cost / max(new_custs, 1)

            # CAC Payback = CAC / (ARPC * Gross Margin)
            gm = row["tech_gross_margin_pct"] or 0.4
            if isinstance(gm, float) and gm < 1:
                monthly_gp_per_cust = arpc * gm
            else:
                monthly_gp_per_cust = arpc * (gm / 100)
            row["cac_payback_months"] = row["cac"] / monthly_gp_per_cust if monthly_gp_per_cust > 0 else None

            # LTV = ARPC * 12 * GM / churn_rate
            churn_rate = row["revenue_churn_pct"] or 0.02
            if isinstance(churn_rate, float) and churn_rate < 1:
                annual_churn = churn_rate * 12
            else:
                annual_churn = churn_rate
            row["ltv"] = (arpc * 12 * gm) / annual_churn if annual_churn > 0 else None
            row["ltv_cac_ratio"] = row["ltv"] / row["cac"] if row.get("cac") and row["cac"] > 0 else None

    # ─────────────────────────────────────────────
    # 10. REVENUE WATERFALL BRIDGE
    # ─────────────────────────────────────────────
    # Worksheet has a trailing space based on audit
    ws_name = "Revenue Waterfall" if "Revenue Waterfall" in wb.sheetnames else "Revenue Waterfall "
    if ws_name in wb.sheetnames:
        ws_wf = wb[ws_name]
        row["wf_revenue_start"] = safe_float(ws_wf.cell(row=2, column=2).value)
        row["wf_one_off_prev"] = safe_float(ws_wf.cell(row=3, column=2).value)
        row["wf_one_off_ytd"] = safe_float(ws_wf.cell(row=4, column=2).value)
        row["wf_recurring_growth"] = safe_float(ws_wf.cell(row=5, column=2).value)
        row["wf_arr_ytg"] = safe_float(ws_wf.cell(row=6, column=2).value)
        row["wf_weighted_pipeline"] = safe_float(ws_wf.cell(row=7, column=2).value)
        row["wf_budget_assumptions"] = safe_float(ws_wf.cell(row=8, column=2).value)
        row["wf_revenue_gap"] = safe_float(ws_wf.cell(row=9, column=2).value)
        row["wf_revenue_end"] = safe_float(ws_wf.cell(row=10, column=2).value)

    # ─────────────────────────────────────────────
    # CLEAN: Convert None strings, ensure types
    # ─────────────────────────────────────────────
    # Convert percentage fields stored as decimals to percentages
    pct_fields_as_decimal = [
        "tech_gross_margin_pct", "tech_gross_margin_ytd_pct",
        "tech_gross_margin_prior_pct", "tech_gross_margin_budget_pct",
        "ebitda_margin_pct", "ebitda_margin_ytd_pct",
        "ebitda_margin_budget_pct", "ebitda_margin_prior_pct",
        "payroll_pct_revenue",
    ]
    for f in pct_fields_as_decimal:
        v = row.get(f)
        if v is not None and isinstance(v, float) and -1 < v < 1:
            row[f] = v * 100

    return row


def parse_and_load(file_path, portco_id="portco-alpha", project_id="averroes-portfolio-intel"):
    """Parse MA file and insert into BigQuery gold.kpi_monthly."""
    from google.cloud import bigquery

    kpi_row = parse_ma_file(file_path, portco_id)

    client = bigquery.Client(project=project_id)
    table_ref = f"{project_id}.gold.kpi_monthly"

    # Remove None values for BQ insertion
    clean_row = {k: v for k, v in kpi_row.items() if v is not None}

    errors = client.insert_rows_json(table_ref, [clean_row])
    if errors:
        print(f"BigQuery insert errors: {errors}")
        return None
    else:
        print(f"Inserted 1 row for {portco_id} period {kpi_row['period']}")
        return kpi_row


if __name__ == "__main__":
    import json
    import sys

    # Local test: parse and print
    file_path = sys.argv[1] if len(sys.argv) > 1 else "MAfileFeb26.xlsx"
    result = parse_ma_file(file_path)
    print(json.dumps(result, indent=2, default=str))
