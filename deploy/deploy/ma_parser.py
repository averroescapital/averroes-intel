import io
import re
import uuid
import openpyxl
import pandas as pd
from datetime import datetime
from parsers.alpha_parser import parse_alpha_ma

def ingest_ma_to_bronze(file_content, file_name, portco_id):
    """
    Main Entry Point for Management Account Parsing.
    Routes to specialized sub-parsers for complex PortCo requirements.
    """
    ingestion_id = str(uuid.uuid4())
    rows_to_insert = []
    
    # 1. SPECIALIZED ROUTING (Portco Alpha - Hospitality SaaS)
    if any(id_match in portco_id.lower() for id_match in ["alpha", "hotel", "hospitality"]):
        print(f"Routing to Specialized Alpha Parser for: {portco_id}")
        alpha_rows = parse_alpha_ma(file_content, file_name)
        for row in alpha_rows:
            row.update({
                "ingestion_id": ingestion_id,
                "portco_id": portco_id,
                "file_name": file_name,
                "value_type": "actual"
            })
        return alpha_rows

    # 2. STANDARD PARSER (Core PE Metrics)
    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    # [Rest of the existing standard parsing logic follows here...]
    # (Existing standard logic remains as fallback for Beta/Gamma)
    col_mappings = {
        2: 'actual', # B
        3: 'budget', # C
        5: 'prior_year', # E
    }
    
    for sheet_name in pnl_sheets:
        if sheet_name not in wb.sheetnames: continue
        ws = wb[sheet_name]
        
        # Period from B3
        period_val = ws.cell(row=3, column=2).value
        reporting_period = parse_date(period_val)
        
        for r in range(4, ws.max_row + 1):
            row_label = ws.cell(row=r, column=1).value
            if not row_label or str(row_label).strip() == "": continue
            
            label = str(row_label).strip().lower()
            # Standardize labels for Gold layer (Indestructible mapping)
            clean_label = None
            if any(x in label for x in ['total revenue', 'revenues', 'turnover', 'sales', 'total income']): clean_label = "TOTAL_REVENUE"
            elif any(x in label for x in ['gross profit', 'gp margin']): clean_label = "GROSS_PROFIT"
            elif any(x in label for x in ['total direct contribution', 'contribution margin']): clean_label = "DIRECT_CONTRIBUTION"
            elif label == 'gross margin': clean_label = "GROSS_MARGIN_PCT"
            elif 'ebitda' in label and 'margin' not in label: clean_label = "ADJUSTED_EBITDA"
            
            if clean_label:
                for c, v_type in col_mappings.items():
                    val = ws.cell(row=r, column=c).value
                    if isinstance(val, (int, float)):
                        rows_to_insert.append({
                            "ingestion_id": ingestion_id,
                            "portco_id": portco_id,
                            "file_name": file_name,
                            "sheet_name": sheet_name,
                            "reporting_period": reporting_period,
                            "row_label": clean_label,
                            "column_label": v_type,
                            "value": float(val),
                            "value_type": v_type,
                            "source_cell": f"{ws.cell(row=r, column=c).coordinate}"
                        })

    # 2. HEADCOUNT SHEET
    if 'Headcount' in wb.sheetnames:
        ws = wb['Headcount']
        period_val = ws.cell(row=3, column=3).value # B3/C3
        reporting_period = parse_date(period_val)
        
        hc_cols = {3: 'actual', 4: 'budget', 6: 'prior_year'} # C, D, F
        for r in range(4, ws.max_row + 1):
            row_label = ws.cell(row=r, column=2).value # Team in B
            if not row_label: continue
            
            for c, v_type in hc_cols.items():
                val = ws.cell(row=r, column=c).value
                if isinstance(val, (int, float)):
                    rows_to_insert.append({
                        "ingestion_id": ingestion_id,
                        "portco_id": portco_id,
                        "file_name": file_name,
                        "sheet_name": 'Headcount',
                        "reporting_period": reporting_period,
                        "row_label": str(row_label).strip(),
                        "column_label": v_type,
                        "value": float(val),
                        "value_type": v_type,
                        "source_cell": f"{ws.cell(row=r, column=c).coordinate}"
                    })

    # 3. BALANCE SHEET SHEET
    if 'Balance Sheet' in wb.sheetnames:
        ws = wb['Balance Sheet']
        period_val = ws.cell(row=3, column=2).value # B3
        reporting_period = parse_date(period_val)
        
        for r in range(4, ws.max_row + 1):
            row_label = ws.cell(row=r, column=1).value # Label in A
            if not row_label: continue
            
            label_str = str(row_label).strip().lower()
            if 'cash at bank' in label_str or 'cash and cash equivalents' in label_str:
                val = ws.cell(row=r, column=2).value # Actual in B
                if isinstance(val, (int, float)):
                    rows_to_insert.append({
                        "ingestion_id": ingestion_id,
                        "portco_id": portco_id,
                        "file_name": file_name,
                        "sheet_name": 'Balance Sheet',
                        "reporting_period": reporting_period,
                        "row_label": "CASH_AT_BANK",
                        "column_label": "actual",
                        "value": float(val),
                        "value_type": "actual",
                        "source_cell": f"{ws.cell(row=r, column=2).coordinate}"
                    })

    # 4. FINANCIAL KPIs SHEET (Block layout)
    if 'Financial KPIs' in wb.sheetnames:
        ws = wb['Financial KPIs']
        reporting_period = parse_date(ws.cell(row=1, column=4).value) # D1
        
        # Financial KPIs extraction (Authoritative cell mapping from PE Fund request)
        kpi_map = [
            (22, 3, "ARPC"), # R22 C
            (22, 11, "YTD_REVENUE_GROWTH"), # R22 K
            (22, 14, "SM_EFFICIENCY"), # R22 N
            (38, 3, "TECH_GROSS_MARGIN_MONTH"), # R38 C
            (38, 11, "EBITDA_MARGIN_MONTH"), # R38 K
            (54, 3, "NET_WORKING_CAPITAL"), # R54 C
            (54, 11, "FREE_CASH_CONVERSION"), # R54 K
            (70, 11, "REVENUE_CHURN_PCT"), # R70 K
            (70, 15, "TIME_TO_VALUE_DAYS"), # R70 O
            (69, 3, "INDICATIVE_EV"), # R69 C
        ]
        
        for r, c, label in kpi_map:
            val = ws.cell(row=r, column=c).value
            if isinstance(val, (int, float)):
                rows_to_insert.append({
                    "ingestion_id": ingestion_id,
                    "portco_id": portco_id,
                    "file_name": file_name,
                    "sheet_name": "Financial KPIs",
                    "reporting_period": reporting_period,
                    "row_label": label,
                    "column_label": "actual",
                    "value": float(val),
                    "value_type": "actual",
                    "source_cell": f"{ws.cell(row=r, column=c).coordinate}"
                })

        # Rule of 40 calculation (Growth + EBITDA Margin)
        # We'll calculate this in Gold view transformation, but we'll extract components here if needed
        # Or if available in a specific cell (not shown in image)

    return rows_to_insert
