import openpyxl
wb = openpyxl.load_workbook('local_MAfileFeb26.xlsx', data_only=True)
# The sheet name has a trailing space
ws = wb['P&L Summary ']
print("Audit of 'P&L Summary ' - ACTUAL Column (B):")
for r in range(1, 40):
    label = ws.cell(row=r, column=1).value
    actual = ws.cell(row=r, column=2).value
    if label and ('Cost' in str(label) or 'Gross' in str(label) or 'Contribution' in str(label)):
        print(f"R{r:02d}: {label} = {actual}")
