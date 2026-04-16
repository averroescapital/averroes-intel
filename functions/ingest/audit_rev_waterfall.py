import openpyxl
wb = openpyxl.load_workbook('local_MAfileFeb26.xlsx', data_only=True)
if 'Revenue Waterfall' in wb.sheetnames:
    ws = wb['Revenue Waterfall']
    print("Revenue Waterfall Sheet Audit:")
    for r in range(1, 15):
        print(f"R{r}: {[ws.cell(row=r, column=c).value for c in range(1, 8)]}")
else:
    print("Sheet 'Revenue Waterfall' not found.")
