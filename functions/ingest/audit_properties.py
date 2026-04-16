import openpyxl
from datetime import datetime
wb = openpyxl.load_workbook('local_MAfileFeb26.xlsx', data_only=True)
ws_cust = wb["Customer Numbers"]
print("Audit of 'Customer Numbers' sheet more columns:")

target_col = None
for c in range(1, 50):
    val = ws_cust.cell(row=2, column=c).value
    if isinstance(val, datetime):
        print(f"Col {c}: {val.strftime('%Y-%m')}")
        if val.year == 2026 and val.month == 2:
            target_col = c
            break

if target_col:
    print(f"Found Feb 2026 at Column {target_col}")
    for r in range(1, 15):
        label = ws_cust.cell(row=r, column=1).value
        val = ws_cust.cell(row=r, column=target_col).value
        print(f"R{r}: Label='{label}' | Value={val}")
else:
    print("Could not find Feb 2026 column.")
