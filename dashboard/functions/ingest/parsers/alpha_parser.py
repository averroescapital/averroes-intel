import io
import openpyxl
from datetime import datetime


def parse_date(val):
    """Normalise any date value to first-of-month ISO string."""
    if isinstance(val, datetime):
        return val.replace(day=1).strftime('%Y-%m-%d')
    if hasattr(val, 'year'):          # date object
        return val.replace(day=1).strftime('%Y-%m-%d')
    return None


def _row(period, label, value, typ, sheet):
    return {"period": period, "kpi": label, "value": float(value), "type": typ, "sheet": sheet}


def parse_alpha_ma(file_content, file_name):
    """
    Production parser for Portco Alpha (Hospitality SaaS).
    Reads P&L Detail, Financial KPIs, Balance Sheet, Headcount, Revenue Waterfall.
    Returns rows ready for bronze normalisation via ma_parser.ingest_ma_to_bronze().
    Each row: { period, kpi (=row_label), value, type (=column_label), sheet }
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    rows = []

    # ------------------------------------------------------------------
    # 1. P&L DETAIL
    #    Col 2 = Actual, Col 3 = Budget, Col 5 = Prior Year, Col 7 = YTD Actual
    #    Period in Row 3, Col 2
    # ------------------------------------------------------------------
    if 'P&L Detail' in wb.sheetnames:
        ws = wb['P&L Detail']
        period = parse_date(ws.cell(3, 2).value)

        # Scan all rows and match by exact label
        label_map = {
            'Ecommerce':                ('ECOMMERCE_REVENUE', True),   # (label, has_budget)
            'EMS':                      ('EMS_REVENUE',       True),
            'Services':                 ('SERVICES_REVENUE',  True),
            'Total Revenue':            ('TOTAL_REVENUE',     True),
            'Total Direct Costs':       ('TOTAL_DIRECT_COSTS', False),
            'Total Direct Contribution':('TOTAL_DIRECT_CONTRIBUTION', True),
            'Total Overheads':          ('TOTAL_OVERHEADS',   False),
            'EBITDA':                   ('EBITDA',            True),
        }

        for r in range(1, ws.max_row + 1):
            cell_label = str(ws.cell(r, 1).value or '').strip()
            if cell_label in label_map:
                bq_label, capture_budget = label_map[cell_label]
                actual = ws.cell(r, 2).value
                budget = ws.cell(r, 3).value
                prior  = ws.cell(r, 5).value

                if isinstance(actual, (int, float)):
                    rows.append(_row(period, bq_label, actual, 'actual', 'P&L Detail'))
                if capture_budget and isinstance(budget, (int, float)):
                    rows.append(_row(period, bq_label, budget, 'budget', 'P&L Detail'))
                if isinstance(prior, (int, float)):
                    rows.append(_row(period, bq_label, prior, 'prior_year', 'P&L Detail'))

        # YTD Revenue (col 7, row 8 = Total Revenue)
        ytd_rev = ws.cell(8, 7).value
        if isinstance(ytd_rev, (int, float)):
            rows.append(_row(period, 'TOTAL_REVENUE_YTD', ytd_rev, 'actual', 'P&L Detail'))

    # ------------------------------------------------------------------
    # 2. FINANCIAL KPIs
    #    Period in Row 1, Col 4
    #    Structure: label in col B, row below has actual in col C
    # ------------------------------------------------------------------
    if 'Financial KPIs' in wb.sheetnames:
        ws = wb['Financial KPIs']
        period = parse_date(ws.cell(1, 4).value)

        # (row, col, label, type)
        kpi_cells = [
            (4,  3, 'TECH_MRR', 'prior_year'),
            (5,  3, 'TECH_MRR', 'budget'),
            (6,  3, 'TECH_MRR', 'actual'),
            (6,  4, 'TECH_MRR_VS_BUDGET_PCT', 'actual'),
            (21, 3, 'ARPC', 'budget'),
            (22, 3, 'ARPC', 'actual'),
            (36, 3, 'TECH_GROSS_MARGIN_MONTH', 'prior_year'),
            (37, 3, 'TECH_GROSS_MARGIN_MONTH', 'budget'),
            (38, 3, 'TECH_GROSS_MARGIN_MONTH', 'actual'),
            (52, 3, 'NET_WORKING_CAPITAL', 'prior_year'),
            (53, 3, 'NET_WORKING_CAPITAL', 'budget'),
            (54, 3, 'NET_WORKING_CAPITAL', 'actual'),
        ]

        for r, c, label, typ in kpi_cells:
            val = ws.cell(r, c).value
            if isinstance(val, (int, float)):
                rows.append(_row(period, label, val, typ, 'Financial KPIs'))

    # ------------------------------------------------------------------
    # 3. BALANCE SHEET
    #    Period in Row 3, Col 3 (Actual column)
    #    Cash is in col 3 (Actual)
    # ------------------------------------------------------------------
    if 'Balance Sheet' in wb.sheetnames:
        ws = wb['Balance Sheet']
        period = parse_date(ws.cell(3, 3).value)

        for r in range(1, min(200, ws.max_row + 1)):
            label = str(ws.cell(r, 2).value or '').strip().lower()
            if label == 'cash':
                val = ws.cell(r, 3).value   # Actual column
                bud = ws.cell(r, 4).value   # Budget column
                if isinstance(val, (int, float)):
                    rows.append(_row(period, 'CASH_BALANCE', val, 'actual', 'Balance Sheet'))
                if isinstance(bud, (int, float)):
                    rows.append(_row(period, 'CASH_BALANCE', bud, 'budget', 'Balance Sheet'))
                break

    # ------------------------------------------------------------------
    # 4. HEADCOUNT
    #    Period in Row 3, Col 3
    #    Team in Col 2, Actual in Col 3, Budget in Col 4
    # ------------------------------------------------------------------
    if 'Headcount' in wb.sheetnames:
        ws = wb['Headcount']
        period = parse_date(ws.cell(3, 3).value)

        for r in range(4, ws.max_row + 1):
            team = ws.cell(r, 2).value
            if not team or not str(team).strip():
                continue
            team_label = str(team).strip()
            actual = ws.cell(r, 3).value
            budget = ws.cell(r, 4).value
            if isinstance(actual, (int, float)):
                rows.append(_row(period, team_label, actual, 'actual', 'Headcount'))
            if isinstance(budget, (int, float)):
                rows.append(_row(period, team_label, budget, 'budget', 'Headcount'))

    # ------------------------------------------------------------------
    # 5. REVENUE WATERFALL
    #    Fixed rows, col 2 = value
    # ------------------------------------------------------------------
    if 'Revenue Waterfall' in wb.sheetnames:
        ws = wb['Revenue Waterfall']
        # Use P&L Detail period (waterfall doesn't have its own date cell)
        period_ws = wb.get('P&L Detail') or wb.get('P&L Summary ')
        period = parse_date(period_ws.cell(3, 2).value) if period_ws else None

        waterfall_map = [
            (2,  'WF_REVENUE_START'),
            (3,  'WF_ONE_OFF_PREV'),
            (4,  'WF_ONE_OFF_YTD'),
            (5,  'WF_RECURRING_GROWTH'),
            (6,  'WF_ARR_TO_GO'),
            (7,  'WF_WEIGHTED_PIPELINE'),
            (8,  'WF_BUDGET_ASSUMPTIONS'),
            (10, 'WF_REVENUE_END'),
        ]
        for r, label in waterfall_map:
            val = ws.cell(r, 2).value
            if isinstance(val, (int, float)) and period:
                rows.append(_row(period, label, val, 'actual', 'Revenue Waterfall'))

    return rows
